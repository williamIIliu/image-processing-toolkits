#!/usr/bin/env python3
"""
Image Processing MCP Server

This server provides comprehensive media processing capabilities including:
- Basic image operations: crop, rotate, resize, flip
- Color adjustments: brightness, contrast, saturation
- Filters and effects
- Format conversion
- Metadata extraction
- Thumbnail generation
- Watermark addition
- file2image utilities: convert WORD, PPTX, and EXCEL to images
- Video utilities: metadata inspection, per-second frame sampling, per-second burst capture,
  single-second snapshots, and frame-by-index exports

The server uses PIL (Pillow) for image processing and includes memory management
and caching for efficient operation.
"""

from __future__ import annotations
from typing import Any, Dict, List, Optional, Union
import os
import subprocess
import threading
import time
import psutil
import tempfile
import shutil
import io
import json
import math
from pathlib import Path

from PIL import Image, ImageEnhance, ImageFilter, ImageDraw, ImageFont, ImageChops
from PIL.ExifTags import TAGS
from mcp.server.fastmcp import FastMCP
from pptxtoimages.tools import PPTXToImageConverter
from pdf2image import convert_from_path
import fitz 
import imageio
from moviepy.video.io.VideoFileClip import VideoFileClip

# Try to import openpyxl
import openpyxl
from openpyxl.utils import get_column_letter

def _ensure_video_path(video_path: str) -> Path:
    """Validate the incoming video path and return it as a Path object."""
    path = Path(video_path)
    if not path.exists():
        raise ImageProcessingError(f"Video not found at: {video_path}")
    return path

def _save_frame_to_path(frame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    imageio.imwrite(output_path, frame)

class ImageProcessorError(Exception):
    """Base class for image processing related errors"""
    pass

class ImageLoadError(ImageProcessorError):
    """Image loading error"""
    pass

class ImageProcessingError(ImageProcessorError):
    """Image processing error"""
    pass

class ImageManager:
    """Image processing manager with caching and memory management"""
    _instance = None
    _lock = threading.Lock()
    _cache = {}
    _last_used = {}
    _memory_threshold = 0.8  # Memory usage threshold
    _max_cache_size = 50  # Maximum number of cached images
    _max_idle_time = 300  # Maximum cache idle time (seconds)

    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        if hasattr(self, '_initialized'):
            return
        self._initialized = True
        self._cleanup_thread = threading.Thread(target=self._cleanup_cache, daemon=True)
        self._cleanup_thread.start()

    def load_image(self, file_path: str) -> Image.Image:
        """Load image with caching support"""
        with self._lock:
            # Check cache
            if file_path in self._cache:
                self._last_used[file_path] = time.time()
                return self._cache[file_path].copy()
            
            # Check memory usage
            self._check_memory_usage()
            
            try:
                # Load image
                image = Image.open(file_path)
                # Convert to RGB mode for compatibility
                if image.mode != 'RGB':
                    image = image.convert('RGB')
                
                # Cache image
                if len(self._cache) < self._max_cache_size:
                    self._cache[file_path] = image.copy()
                    self._last_used[file_path] = time.time()
                
                return image
            except Exception as e:
                raise ImageLoadError(f"Unable to load image {file_path}: {str(e)}")

    def _check_memory_usage(self):
        """Check memory usage and clean cache if necessary"""
        memory_percent = psutil.virtual_memory().percent / 100
        if memory_percent > self._memory_threshold:
            self._clear_oldest_cache()

    def _clear_oldest_cache(self):
        """Clear the oldest cache entry"""
        if not self._last_used:
            return
        
        oldest_key = min(self._last_used.items(), key=lambda x: x[1])[0]
        if oldest_key in self._cache:
            del self._cache[oldest_key]
            del self._last_used[oldest_key]

    def _cleanup_cache(self):
        """Periodically clean expired cache"""
        while True:
            time.sleep(60)  # Check every minute
            with self._lock:
                current_time = time.time()
                expired_keys = [
                    key for key, last_used in self._last_used.items()
                    if current_time - last_used > self._max_idle_time
                ]
                for key in expired_keys:
                    if key in self._cache:
                        del self._cache[key]
                    del self._last_used[key]

    def clear_cache(self):
        """Clear all cache"""
        with self._lock:
            self._cache.clear()
            self._last_used.clear()

# Create global ImageManager instance
image_manager = ImageManager()

# Create FastMCP server
mcp = FastMCP("Image-Processing-Toolkits")

@mcp.tool()
def crop_image(file_path: str, ymin: int, xmin: int, ymax: int, xmax: int, output_path: Optional[str] = None) -> str:
    """Crop the image to a specified rectangular region defined by normalized bounding box coordinates (0-1000).
    
    This tool is designed to work directly with Vision Language Model outputs which typically provide
    coordinates in [ymin, xmin, ymax, xmax] format on a 0-1000 scale.

    Args:
        file_path: Path to the source image file.
        ymin: Top edge normalized coordinate (0-1000).
        xmin: Left edge normalized coordinate (0-1000).
        ymax: Bottom edge normalized coordinate (0-1000).
        xmax: Right edge normalized coordinate (0-1000).
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_cropped<ext>" will be created next to
            the source image.

    Returns:
        A human-readable message describing where the cropped image is saved,
        or an error message if the requested region exceeds image boundaries.
    """
    try:
        image = image_manager.load_image(file_path)
        img_width, img_height = image.size

        # Convert normalized coordinates (0-1000) to absolute pixels
        # Formula: pixel = (normalized / 1000) * dimension
        x0 = int((xmin / 1000) * img_width)
        y0 = int((ymin / 1000) * img_height)
        x1 = int((xmax / 1000) * img_width)
        y1 = int((ymax / 1000) * img_height)
        
        # Ensure correct coordinates order
        x_min, x_max = min(x0, x1), max(x0, x1)
        y_min, y_max = min(y0, y1), max(y0, y1)
        
        # Validate crop region
        if x_min < 0 or y_min < 0 or x_max > img_width or y_max > img_height:
            return f"Crop region exceeds image boundaries. Image size: {img_width}x{img_height}, Crop: [{x_min}, {y_min}, {x_max}, {y_max}]"
        
        # Crop image
        cropped = image.crop((x_min, y_min, x_max, y_max))
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_cropped{ext}"
        
        cropped.save(output_path)
        return f"Image cropped and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error cropping image: {str(e)}")

@mcp.tool()
def apply_mosaic(file_path: str, x0: int, y0: int, x1: int, y1: int,
                 block_size: int = 16, output_path: Optional[str] = None) -> str:
    """Apply a mosaic (pixelation) effect to a rectangular region defined by bounding box coordinates.
    This is typically used for anonymization, e.g. masking faces, license
    plates, or other sensitive areas while leaving the rest of the image
    unchanged.

    Args:
        file_path: Path to the source image file.
        x0: X-coordinate of the top-left corner (pixels).
        y0: Y-coordinate of the top-left corner (pixels).
        x1: X-coordinate of the bottom-right corner (pixels).
        y1: Y-coordinate of the bottom-right corner (pixels).
        block_size: Size of mosaic blocks. Larger values produce coarser
            pixelation. Must be a positive integer.
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_mosaic_x0_y0_x1_y1<ext>" will be
            created next to the source image.

    Returns:
        A human-readable message describing where the mosaic image is saved,
        or an error message if the region is out of bounds or parameters are
        invalid.
    """
    try:
        image = image_manager.load_image(file_path)
        # Ensure correct coordinates
        x_min, x_max = min(x0, x1), max(x0, x1)
        y_min, y_max = min(y0, y1), max(y0, y1)
        width = x_max - x_min
        height = y_max - y_min
        img_width, img_height = image.size

        if x_min < 0 or y_min < 0 or x_max > img_width or y_max > img_height:
            return f"Mosaic region exceeds image boundaries. Image size: {img_width}x{img_height}"

        if block_size <= 0:
            return "block_size must be a positive integer"

        # Extract region to pixelate
        region = image.crop((x_min, y_min, x_max, y_max))
        
        # Compute downscaled size for pixelation (at least 1x1)
        small_w = max(1, width // block_size)
        small_h = max(1, height // block_size)

        # Downscale then upscale using NEAREST to create mosaic blocks
        mosaic_small = region.resize((small_w, small_h), Image.Resampling.NEAREST)
        mosaic_region = mosaic_small.resize((width, height), Image.Resampling.NEAREST)

        # Paste back into original image
        image.paste(mosaic_region, (x_min, y_min))

        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_mosaic_{x_min}_{y_min}_{x_max}_{y_max}{ext}"
        
        image.save(output_path)
        return f"Mosaic applied to region and saved to: {output_path}"

    except Exception as e:
        raise ImageProcessingError(f"Error applying mosaic: {str(e)}")

@mcp.tool()
def rotate_image(file_path: str, angle: float, expand: bool = True, output_path: Optional[str] = None) -> str:
    """Rotate the image by a specified angle.

    Args:
        file_path: Path to the source image file.
        angle: Rotation angle in degrees. Positive values rotate counter-
            clockwise; negative values rotate clockwise.
        expand: Whether to expand the canvas to fit the entire rotated image.
            If False, parts of the image may be cropped.
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_rotated_<angle>deg<ext>" will be created.

    Returns:
        A human-readable message describing where the rotated image is saved.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Rotate image
        rotated = image.rotate(angle, expand=expand, fillcolor='white')
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_rotated_{angle}deg{ext}"
        
        rotated.save(output_path)
        return f"Image rotated {angle} degrees and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error rotating image: {str(e)}")

@mcp.tool()
def resize_image(file_path: str, width: int, height: int, maintain_aspect: bool = True, 
                resample: str = "LANCZOS", output_path: Optional[str] = None) -> str:
    """Resize the image to the specified dimensions.

    Args:
        file_path: Path to the source image file.
        width: Target width in pixels.
        height: Target height in pixels.
        maintain_aspect: If True, preserve the original aspect ratio. The
            resulting image will fit within (width, height). If False, the
            image is stretched exactly to (width, height).
        resample: Resampling method to use. One of
            "LANCZOS", "BILINEAR", "BICUBIC", "NEAREST".
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_resized_<width>x<height><ext>" will be
            created.

    Returns:
        A human-readable message describing where the resized image is saved.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Get resampling method
        resample_methods = {
            "LANCZOS": Image.Resampling.LANCZOS,
            "BILINEAR": Image.Resampling.BILINEAR,
            "BICUBIC": Image.Resampling.BICUBIC,
            "NEAREST": Image.Resampling.NEAREST,
        }
        resample_method = resample_methods.get(resample, Image.Resampling.LANCZOS)
        
        if maintain_aspect:
            # Maintain aspect ratio
            image.thumbnail((width, height), resample_method)
            resized = image
        else:
            # Force resize to specified dimensions
            resized = image.resize((width, height), resample_method)
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_resized_{width}x{height}{ext}"
        
        resized.save(output_path)
        return f"Image resized and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error resizing image: {str(e)}")

@mcp.tool()
def adjust_contrast(file_path: str, factor: float, output_path: Optional[str] = None) -> str:
    """Adjust the global contrast of an image.

    Args:
        file_path: Path to the source image file.
        factor: Contrast factor. 1.0 keeps the original contrast; values
            greater than 1.0 increase contrast; values between 0 and 1.0
            reduce contrast.
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_contrast_<factor><ext>" will be created.

    Returns:
        A human-readable message describing where the adjusted image is saved.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Adjust contrast
        enhancer = ImageEnhance.Contrast(image)
        enhanced = enhancer.enhance(factor)
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_contrast_{factor}{ext}"
        
        enhanced.save(output_path)
        return f"Image contrast adjusted and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error adjusting contrast: {str(e)}")

@mcp.tool()
def adjust_brightness(file_path: str, factor: float, output_path: Optional[str] = None) -> str:
    """Adjust the global brightness of an image.

    Args:
        file_path: Path to the source image file.
        factor: Brightness factor. 1.0 keeps original brightness; values
            greater than 1.0 make the image brighter; values between 0 and
            1.0 make it darker.
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_brightness_<factor><ext>" will be created.

    Returns:
        A human-readable message describing where the adjusted image is saved.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Adjust brightness
        enhancer = ImageEnhance.Brightness(image)
        enhanced = enhancer.enhance(factor)
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_brightness_{factor}{ext}"
        
        enhanced.save(output_path)
        return f"Image brightness adjusted and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error adjusting brightness: {str(e)}")

@mcp.tool()
def adjust_saturation(file_path: str, factor: float, output_path: Optional[str] = None) -> str:
    """Adjust the color saturation of an image.

    Args:
        file_path: Path to the source image file.
        factor: Saturation factor. 1.0 keeps original colors; values greater
            than 1.0 make colors more vivid; values between 0 and 1.0 reduce
            saturation; 0 turns the image grayscale.
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_saturation_<factor><ext>" will be created.

    Returns:
        A human-readable message describing where the adjusted image is saved.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Adjust saturation
        enhancer = ImageEnhance.Color(image)
        enhanced = enhancer.enhance(factor)
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_saturation_{factor}{ext}"
        
        enhanced.save(output_path)
        return f"Image saturation adjusted and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error adjusting saturation: {str(e)}")

@mcp.tool()
def apply_filter(file_path: str, filter_type: str, output_path: Optional[str] = None) -> str:
    """Apply a built-in Pillow filter to the image.

    Supported filter types include:
    BLUR, DETAIL, EDGE_ENHANCE, EDGE_ENHANCE_MORE, EMBOSS,
    FIND_EDGES, SHARPEN, SMOOTH, SMOOTH_MORE.

    Args:
        file_path: Path to the source image file.
        filter_type: Name of the filter to apply (case-sensitive, must be one
            of the supported filter names).
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_<filter_type.lower()><ext>" will be
            created.

    Returns:
        A human-readable message describing where the filtered image is saved,
        or an error message if the filter type is unknown.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Get filter
        filters = {
            "BLUR": ImageFilter.BLUR,
            "DETAIL": ImageFilter.DETAIL,
            "EDGE_ENHANCE": ImageFilter.EDGE_ENHANCE,
            "EDGE_ENHANCE_MORE": ImageFilter.EDGE_ENHANCE_MORE,
            "EMBOSS": ImageFilter.EMBOSS,
            "FIND_EDGES": ImageFilter.FIND_EDGES,
            "SHARPEN": ImageFilter.SHARPEN,
            "SMOOTH": ImageFilter.SMOOTH,
            "SMOOTH_MORE": ImageFilter.SMOOTH_MORE,
        }
        
        filter_obj = filters.get(filter_type)
        if not filter_obj:
            return f"Unknown filter type: {filter_type}"
        
        # Apply filter
        filtered = image.filter(filter_obj)
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_{filter_type.lower()}{ext}"
        
        filtered.save(output_path)
        return f"Applied {filter_type} filter and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error applying filter: {str(e)}")

@mcp.tool()
def flip_image(file_path: str, direction: str, output_path: Optional[str] = None) -> str:
    """Flip the image horizontally or vertically.

    Args:
        file_path: Path to the source image file.
        direction: Flip direction. Use "horizontal" for a left-right mirror
            flip, or "vertical" for a top-bottom flip.
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_flip_<direction><ext>" will be created.

    Returns:
        A human-readable message describing where the flipped image is saved,
        or an error message if the direction is invalid.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Flip image
        if direction == "horizontal":
            flipped = image.transpose(Image.Transpose.FLIP_LEFT_RIGHT)
        elif direction == "vertical":
            flipped = image.transpose(Image.Transpose.FLIP_TOP_BOTTOM)
        else:
            return f"Unknown flip direction: {direction}"
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_flip_{direction}{ext}"
        
        flipped.save(output_path)
        return f"Image flipped {direction} and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error flipping image: {str(e)}")

@mcp.tool()
def convert_format(file_path: str, format: str, quality: int = 95, output_path: Optional[str] = None) -> str:
    """Convert the image to a different file format.

    Args:
        file_path: Path to the source image file.
        format: Target format name (e.g. "JPEG", "PNG", "BMP", "TIFF", "WEBP").
        quality: JPEG quality setting (1-100). Ignored for non-JPEG formats.
        output_path: Optional path for the output image. If not provided,
            a new file will be created next to the source image with the
            appropriate extension.

    Returns:
        A human-readable message describing the new format and output path.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Generate output path
        if not output_path:
            name = os.path.splitext(file_path)[0]
            ext = format.lower()
            if ext == "jpeg":
                ext = "jpg"
            output_path = f"{name}.{ext}"
        
        # Save in new format
        save_kwargs = {}
        if format.upper() == "JPEG":
            save_kwargs["quality"] = quality
            save_kwargs["optimize"] = True
        
        image.save(output_path, format=format.upper(), **save_kwargs)
        return f"Image converted to {format} format and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error converting format: {str(e)}")

@mcp.tool()
def get_image_info(file_path: str) -> str:
    """Return human-readable information and metadata for an image.

    The returned string includes basic properties (format, mode, width,
    height, file size) and, when available, EXIF metadata such as camera
    model, capture time, and other embedded fields.

    Args:
        file_path: Path to the image file.

    Returns:
        A formatted multi-line string describing image properties and EXIF
        metadata, or an error message if the image cannot be read.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Basic information
        info = {
            "File Path": file_path,
            "Format": image.format,
            "Mode": image.mode,
            "Size": f"{image.size[0]}x{image.size[1]}",
            "Width": image.size[0],
            "Height": image.size[1],
        }
        
        # File size
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            info["File Size"] = f"{file_size / 1024:.2f} KB"
        
        # EXIF data
        if hasattr(image, '_getexif') and image._getexif():
            exif_data = {}
            for tag_id, value in image._getexif().items():
                tag = TAGS.get(tag_id, tag_id)
                exif_data[tag] = value
            info["EXIF Data"] = exif_data
        
        # Format output
        text = "Image Information:\n"
        for key, value in info.items():
            if key == "EXIF Data" and isinstance(value, dict):
                text += f"\n{key}:\n"
                for exif_key, exif_value in value.items():
                    text += f"  {exif_key}: {exif_value}\n"
            else:
                text += f"{key}: {value}\n"
        
        return text
        
    except Exception as e:
        return f"Error getting image info: {str(e)}"

@mcp.tool()
def create_thumbnail(file_path: str, size: int = 128, output_path: Optional[str] = None) -> str:
    """Create a thumbnail image with the longest side limited to a given size.

    The thumbnail preserves aspect ratio and is suitable for previews,
    gallery views, or low-resolution representations.

    Args:
        file_path: Path to the source image file.
        size: Maximum size (in pixels) of the longer edge of the thumbnail.
        output_path: Optional path for the thumbnail image. If not provided,
            a file named "<original>_thumbnail_<size><ext>" will be created.

    Returns:
        A human-readable message describing where the thumbnail is saved.
    """
    try:
        image = image_manager.load_image(file_path)
        
        # Create thumbnail
        image.thumbnail((size, size), Image.Resampling.LANCZOS)
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_thumbnail_{size}{ext}"
        
        image.save(output_path)
        return f"Thumbnail created and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error creating thumbnail: {str(e)}")

@mcp.tool()
def add_text_to_photo(file_path: str, watermark_text: str, ymin: int, xmin: int,
                 color: str = "white", font_size: int = 36, opacity: float = 0.5, output_path: Optional[str] = None) -> str:
    """Add semi-transparent text to an image using coordinate placement.

    The text is drawn on a transparent overlay and then composited with the
    original image.

    Args:
        file_path: Path to the source image file.
        watermark_text: Text to render.
        ymin: Normalized Y coordinate (0-1000) of the text anchor.
        xmin: Normalized X coordinate (0-1000) of the text anchor.
        color: Text color (named or hex). Named colors supported: "white",
               "black", "red", "green", "blue", "yellow", "cyan", "magenta".
               Hex format "#RRGGBB" is also supported.
        font_size: Font size in pixels.
        opacity: Opacity in [0.0, 1.0].
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_watermarked<ext>" will be created.

    Notes:
        - Anchor: the provided (ymin, xmin) refers to the TOP-LEFT corner of the text.

    Returns:
        A human-readable message describing where the watermarked image is saved.
    """
    try:
        
        image = image_manager.load_image(file_path)
        img_width, img_height = image.size
        
        # Create watermark layer
        watermark = Image.new('RGBA', image.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(watermark)
        
        # Try to load font from common Linux and Windows paths
        font = None
        candidate_fonts = [
            # Linux common
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
            # Windows common
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/arialbd.ttf",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/msyh.ttc",
        ]
        for fp in candidate_fonts:
            try:
                font = ImageFont.truetype(fp, font_size)
                break
            except Exception:
                continue
        if font is None:
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except Exception:
                font = ImageFont.load_default()
        
        # Get text dimensions
        bbox = draw.textbbox((0, 0), watermark_text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        
        # Calculate position from normalized coordinates (0-1000)
        x = int((xmin / 1000) * img_width)
        y = int((ymin / 1000) * img_height)
        
        # Simple color parser (named colors and #RRGGBB)
        def _parse_color(c: str) -> tuple:
            named = {
                "white": (255, 255, 255),
                "black": (0, 0, 0),
                "red": (255, 0, 0),
                "green": (0, 255, 0),
                "blue": (0, 0, 255),
                "yellow": (255, 255, 0),
                "cyan": (0, 255, 255),
                "magenta": (255, 0, 255),
            }
            c = c.lower().strip()
            if c in named:
                return named[c]
            if c.startswith('#') and len(c) == 7:
                try:
                    r = int(c[1:3], 16)
                    g = int(c[3:5], 16)
                    b = int(c[5:7], 16)
                    return (r, g, b)
                except Exception:
                    return named["white"]
            return named["white"]
        
        # Draw watermark
        alpha = int(255 * max(0.0, min(1.0, opacity)))
        r, g, b = _parse_color(color)
        draw.text((x, y), watermark_text, font=font, fill=(r, g, b, alpha))
        
        # Merge images
        watermarked = Image.alpha_composite(image.convert('RGBA'), watermark)
        watermarked = watermarked.convert('RGB')
        
        # Save result
        if not output_path:
            name, ext = os.path.splitext(file_path)
            output_path = f"{name}_watermarked{ext}"
        
        watermarked.save(output_path)
        return f"Watermark added and saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error adding watermark: {str(e)}")

@mcp.tool()
def overlay_image_on_photo(
    base_image_path: str,
    overlay_image_path: str,
    ymin: int,
    xmin: int,
    scale: float = 1.0,
    output_path: Optional[str] = None
) -> str:
    """Paste an overlay image onto a base image using coordinate placement.

    The overlay is pasted preserving its alpha channel, with its TOP-LEFT
    corner aligned to the specified (ymin, xmin) position.

    Args:
        base_image_path: Path to the base image.
        overlay_image_path: Path to the overlay image (PNG with transparency recommended).
        ymin: Normalized Y coordinate (0-1000) for the overlay's top-left corner.
        xmin: Normalized X coordinate (0-1000) for the overlay's top-left corner.
        scale: Scale factor applied to the overlay image size (e.g., 0.5 halves the size).
        output_path: Optional path for the output image. If not provided, it will use
            the base image name with an "_overlay" suffix.

    Returns:
        Success message with output path.
    """
    try:
        # Load base image
        base_img = Image.open(base_image_path).convert('RGBA')
        width, height = base_img.size
        
        # Load overlay image
        if not (overlay_image_path and os.path.exists(overlay_image_path)):
            raise ImageProcessingError(f"Overlay image not found: {overlay_image_path}")
        ov = Image.open(overlay_image_path).convert('RGBA')
        
        # Scale overlay image
        if scale != 1.0:
            new_w = max(1, int(ov.width * scale))
            new_h = max(1, int(ov.height * scale))
            ov = ov.resize((new_w, new_h), Image.Resampling.LANCZOS)
        
        # Position from normalized coordinates (0-1000)
        x = int((xmin / 1000) * width)
        y = int((ymin / 1000) * height)
        
        # Paste overlay
        base_img.paste(ov, (x, y), ov)
        result = base_img
        
        # Convert back to RGB if needed
        if result.mode == 'RGBA':
            rgb_result = Image.new('RGB', result.size, (255, 255, 255))
            rgb_result.paste(result, mask=result.split()[3])
            result = rgb_result
        
        # Generate output path if not provided
        if not output_path:
            base_name, ext = os.path.splitext(base_image_path)
            output_path = f"{base_name}_overlay{ext}"
        
        # Save result
        result.save(output_path, quality=95)
        return f"Successfully added image overlay. Saved to: {output_path}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error adding overlay: {str(e)}")
        
@mcp.tool()
def draw_bounding_box_on_photo(image_path: str, ymin: int, xmin: int, ymax: int, xmax: int,
                               color: str = "red", output_path: Optional[str] = None) -> str:
    """Draw a rectangular bounding box on an image using normalized coordinates.

    Args:
        image_path: Path to the image.
        ymin: Normalized top edge (0-1000).
        xmin: Normalized left edge (0-1000).
        ymax: Normalized bottom edge (0-1000).
        xmax: Normalized right edge (0-1000).
        color: Outline color (named or hex). Named colors supported: "white",
               "black", "red", "green", "blue", "yellow", "cyan", "magenta".
               Hex format "#RRGGBB" is also supported. Default is "red".
        output_path: Optional path for the output image. If not provided,
            a file named "<original>_boxed<ext>" will be created.

    Returns:
        A human-readable message describing where the boxed image is saved.
    """
    try:
        image = image_manager.load_image(image_path)
        width, height = image.size
        
        # Convert normalized coordinates (0-1000) to absolute pixels
        x0 = int((xmin / 1000) * width)
        y0 = int((ymin / 1000) * height)
        x1 = int((xmax / 1000) * width)
        y1 = int((ymax / 1000) * height)
        
        # Ensure correct order
        x_min, x_max = min(x0, x1), max(x0, x1)
        y_min, y_max = min(y0, y1), max(y0, y1)
        
        # Create overlay to draw the rectangle
        overlay = Image.new('RGBA', image.size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(overlay)
        line_w = max(1, int(min(width, height) * 0.005))

        # Parse color
        def _parse_color(c: str) -> tuple:
            named = {
                "white": (255, 255, 255),
                "black": (0, 0, 0),
                "red": (255, 0, 0),
                "green": (0, 255, 0),
                "blue": (0, 0, 255),
                "yellow": (255, 255, 0),
                "cyan": (0, 255, 255),
                "magenta": (255, 0, 255),
            }
            c = c.lower().strip()
            if c in named:
                return named[c]
            if c.startswith('#') and len(c) == 7:
                try:
                    r = int(c[1:3], 16)
                    g = int(c[3:5], 16)
                    b = int(c[5:7], 16)
                    return (r, g, b)
                except Exception:
                    return named["red"]
            return named["red"]
        r, g, b = _parse_color(color)

        draw.rectangle([x_min, y_min, x_max, y_max], outline=(r, g, b, 255), width=line_w)
        
        # Composite and save
        composed = Image.alpha_composite(image.convert('RGBA'), overlay)
        result = composed.convert('RGB')
        if not output_path:
            name, ext = os.path.splitext(image_path)
            output_path = f"{name}_boxed{ext}"
        result.save(output_path)
        return f"Bounding box drawn and saved to: {output_path}"
    except Exception as e:
        raise ImageProcessingError(f"Error drawing bounding box: {str(e)}")

@mcp.tool()
def pptx_to_images(pptx_path: str, output_dir: Optional[str] = None) -> str:
    """Convert a PowerPoint (.pptx) file to images.

    Each slide in the presentation will be converted to a separate PNG image.

    Args:
        pptx_path: Path to the source PowerPoint file (.pptx).
        output_dir: Optional directory for output images. If not provided,
            a directory named "<pptx_name>_images" will be created next to
            the source file.

    Returns:
        A human-readable message listing the converted image paths,
        or an error message if conversion fails.
    """
    try:
        # Generate output directory if not provided
        if not output_dir:
            base_name = os.path.splitext(pptx_path)[0]
            output_dir = f"{base_name}_images"
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Initialize converter and convert
        converter = PPTXToImageConverter(pptx_path, output_dir=output_dir)
        images = converter.convert()
        
        return f"Converted {len(images)} slides to images in: {output_dir}\nImages: {images}"
        
    except Exception as e:
        raise ImageProcessingError(f"Error converting PPTX to images: {str(e)}")


@mcp.tool()
def convert_pdf_to_images(
    pdf_path: str,
    output_dir: str,
    pages: Optional[Union[int, List[int], tuple]] = None,
    dpi: int = 100,
    image_format: str = 'PNG',
    prefix: str = 'page'
) -> str:
    """Convert specified PDF pages or page ranges to images.

    Args:
        pdf_path: Path to the PDF file.
        output_dir: Directory path for output images.
        pages: Page range specification:
            - int: Single page number (0-based)
            - List[int]: List of page numbers [0, 2, 4]
            - tuple: Page range (start_page, end_page) inclusive
            - None: All pages
        image_format: Image format, supports 'PNG', 'JPEG', 'BMP' etc., default 'PNG'
        prefix: Output filename prefix, default 'page'

    Returns:
        A human-readable message listing the generated image file paths,
        or an error message if conversion fails.
    """
    if fitz is None:
        raise ImageProcessingError("PyMuPDF not installed. Please install PyMuPDF to use PDF conversion functionality.")
    
    try:
        # Check if PDF file exists
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        # Open PDF file
        pdf_document = fitz.open(pdf_path)
        total_pages = pdf_document.page_count
        
        # Determine pages to convert
        page_indices = []
        
        if pages is None:
            # Convert all pages
            page_indices = list(range(total_pages))
        elif isinstance(pages, int):
            # Single page
            if 0 <= pages < total_pages:
                page_indices = [pages]
            else:
                raise ValueError(f"Page {pages} out of range, PDF total pages: {total_pages}")
        elif isinstance(pages, list):
            # List of pages
            for page_num in pages:
                if 0 <= page_num < total_pages:
                    page_indices.append(page_num)
                else:
                    raise ValueError(f"Page {page_num} out of range, PDF total pages: {total_pages}")
        elif isinstance(pages, tuple) and len(pages) == 2:
            # Page range
            start_page, end_page = pages
            if 0 <= start_page < total_pages and 0 <= end_page < total_pages:
                page_indices = list(range(start_page, end_page + 1))
            else:
                raise ValueError(f"Page range ({start_page}, {end_page}) out of range, PDF total pages: {total_pages}")
        else:
            raise ValueError("pages parameter must be int, List[int], tuple, or None")
        
        # Convert specified pages
        output_files = []
        
        for page_num in page_indices:
            # Get page
            page = pdf_document[page_num]
            
            # Set zoom factor to control resolution
            zoom = dpi / 72.0  # PDF default DPI is 72
            matrix = fitz.Matrix(zoom, zoom)
            
            # Render page to image
            pix = page.get_pixmap(matrix=matrix)
            
            # Convert to PIL Image
            img_data = pix.tobytes(image_format.lower())
            img = Image.open(io.BytesIO(img_data))
            
            # Generate output filename
            output_filename = f"{prefix}_{page_num + 1}.{image_format.lower()}"
            output_path = os.path.join(output_dir, output_filename)
            
            # Save image
            img.save(output_path, format=image_format)
            output_files.append(output_path)
            
            print(f"Converted page {page_num + 1}: {output_path}")
        
        # Close PDF document
        pdf_document.close()
        
        return f"Successfully converted {len(output_files)} pages to images in: {output_dir}\nImages: {output_files}"
        
    except Exception as e:
        if 'pdf_document' in locals():
            pdf_document.close()
        raise ImageProcessingError(f"PDF conversion failed: {str(e)}")


@mcp.tool()
def word_to_images(docx_path: str, output_dir: Optional[str] = None) -> str:
    """Convert a Word document (.docx) to images.

    The document is first converted to PDF using LibreOffice, then each page
    is converted to a separate PNG image. Requires LibreOffice (soffice) to
    be installed on the system.

    Args:
        docx_path: Path to the source Word document (.docx).
        output_dir: Optional directory for output images. If not provided,
            a directory named "<docx_name>_images" will be created next to
            the source file.

    Returns:
        A human-readable message listing the converted image paths,
        or an error message if conversion fails.
    """
    try:
        # Generate output directory if not provided
        if not output_dir:
            base_name = os.path.splitext(docx_path)[0]
            output_dir = f"{base_name}_images"
        
        os.makedirs(output_dir, exist_ok=True)
        
        # Convert docx to pdf using LibreOffice
        subprocess.run([
            "soffice", "--headless", "--convert-to", "pdf",
            "--outdir", output_dir, docx_path
        ], check=True)
        
        # Get the generated pdf file path
        base_name = os.path.splitext(os.path.basename(docx_path))[0]
        pdf_path = os.path.join(output_dir, f"{base_name}.pdf")
        
        # Convert pdf to images
        images = convert_from_path(pdf_path)
        
        image_paths = []
        for i, image in enumerate(images):
            img_path = os.path.join(output_dir, f"{base_name}_page_{i+1}.png")
            image.save(img_path, "PNG")
            image_paths.append(img_path)
        
        # Clean up temporary pdf
        os.remove(pdf_path)
        
        return f"Converted {len(image_paths)} pages to images in: {output_dir}\nImages: {image_paths}"
        
    except subprocess.CalledProcessError as e:
        raise ImageProcessingError(f"LibreOffice conversion failed: {str(e)}. Make sure LibreOffice is installed.")
    except Exception as e:
        raise ImageProcessingError(f"Error converting Word to images: {str(e)}")


@mcp.tool()
def excel_to_image(file_path: str, sheetname: str = "Sheet1", output_path: Optional[str] = None) -> str:
    """Convert an Excel sheet to an image using LibreOffice.
    
    This tool converts a specific Excel sheet into a single image. It automatically adjusts 
    the print area and page settings to ensure the entire content fits on one page (landscape),
    preventing columns from being cut off.

    Dependencies:
        - LibreOffice (soffice command)
        - pdf2image (and poppler-utils)
        - openpyxl (recommended for best results)

    Args:
        file_path: Path to the Excel file (.xlsx, .xls).
        sheetname: Name of the sheet to convert (default: "Sheet1"). 
                   Note: Currently, due to LibreOffice limitations, this might convert the active sheet 
                   or the first sheet if not specified strictly, but the tool attempts to configure the file correctly.
        output_path: Optional path for the output image (e.g., "output.png"). 
                     If not provided, the image will be saved in the same directory as the Excel file 
                     with the same name and a .png extension.

    Returns:
        A human-readable string indicating success and the path to the saved image, 
        or an error message if the operation fails.
    """
    abs_file_path = os.path.abspath(file_path)
    if not os.path.exists(abs_file_path):
        raise ImageProcessingError(f"Input file not found: {abs_file_path}")

    # Determine output path if not provided
    if not output_path:
        base_name = os.path.splitext(abs_file_path)[0]
        output_path = f"{base_name}.png"

    # Preprocessing helper
    def preprocess_excel(path):
        # Create a temp copy to avoid modifying original
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        shutil.copy2(path, temp_path)
        
        try:
            wb = openpyxl.load_workbook(temp_path)
            for sheet in wb.worksheets:
                sheet.print_area = None
                
                # Calculate the dimension
                max_row = sheet.max_row
                max_col = sheet.max_column
                if max_row > 0 and max_col > 0:
                    last_col_letter = get_column_letter(max_col)
                    sheet.print_area = f"A1:{last_col_letter}{max_row}"
                
                # Page setup
                sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
                sheet.page_setup.fitToPage = True
                sheet.page_setup.fitToWidth = 1
                sheet.page_setup.fitToHeight = 0 
                
                # Margins
                sheet.page_margins.left = 0
                sheet.page_margins.right = 0
                sheet.page_margins.top = 0
                sheet.page_margins.bottom = 0
                sheet.page_margins.header = 0
                sheet.page_margins.footer = 0
                
            wb.save(temp_path)
            wb.close()
            return temp_path, True
        except Exception:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return path, False

    # Image trimming helper
    def trim_whitespace(img):
        bg = Image.new(img.mode, img.size, img.getpixel((0,0)))
        diff = ImageChops.difference(img, bg)
        diff = ImageChops.add(diff, diff, 2.0, -100)
        bbox = diff.getbbox()
        if bbox:
            return img.crop(bbox)
        return img

    processed_path, is_temp = preprocess_excel(abs_file_path)
    output_dir = tempfile.mkdtemp()
    
    # LibreOffice output filename guess
    base_name = os.path.splitext(os.path.basename(processed_path))[0]
    pdf_path = os.path.join(output_dir, f"{base_name}.pdf")

    try:
        # LibreOffice conversion
        cmd = [
            "libreoffice",
            "--headless",
            "--convert-to",
            'pdf:calc_pdf_Export:{"SinglePageSheets":{"type":"boolean","value":"true"}}',
            "--outdir",
            output_dir,
            processed_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            raise ImageProcessingError(f"LibreOffice conversion failed: {result.stderr}")

        # Locate PDF
        if not os.path.exists(pdf_path):
            files = [f for f in os.listdir(output_dir) if f.endswith(".pdf")]
            if files:
                pdf_path = os.path.join(output_dir, files[0])
            else:
                raise ImageProcessingError("PDF file was not generated by LibreOffice.")

        # PDF to Image
        images = convert_from_path(pdf_path, dpi=100)
        if images:
            img = images[0]
            img = trim_whitespace(img)
            img.save(output_path)
            return f"Excel sheet converted and saved to: {output_path}"
        else:
            raise ImageProcessingError("No images extracted from the generated PDF.")

    except Exception as e:
        raise ImageProcessingError(f"Error converting Excel to image: {str(e)}")
        
    finally:
        if os.path.exists(output_dir):
            shutil.rmtree(output_dir)
        if is_temp and os.path.exists(processed_path):
            os.remove(processed_path)

@mcp.tool()
def get_video_metadata_json(video_path: str) -> str:
    """Return JSON metadata (duration, frame size, FPS, approximate frames) for a video file."""
    try:
        path = _ensure_video_path(video_path)
        with VideoFileClip(str(path)) as clip:
            fps = clip.fps or 0
            approx_total_frames = int(round(fps * clip.duration)) if fps else None
            data = {
                "duration_seconds": round(clip.duration, 2),
                "frame_size": clip.size,
                "fps": round(fps, 2),
                "approx_total_frames": approx_total_frames,
            }
        return json.dumps(data, ensure_ascii=False, indent=2)
    except Exception as e:
        raise ImageProcessingError(f"Error reading video metadata: {str(e)}")


@mcp.tool()
def export_frames_every_second(
    video_path: str,
    output_dir: str,
    interval_seconds: int = 1,
) -> str:
    """Capture one frame every `interval_seconds` seconds across the video and store them under `output_dir`."""
    try:
        if interval_seconds <= 0:
            raise ImageProcessingError("interval_seconds must be positive.")

        path = _ensure_video_path(video_path)
        output_path = Path(output_dir)
        exported_files: List[str] = []

        with VideoFileClip(str(path)) as clip:
            duration = math.floor(clip.duration)
            output_path.mkdir(parents=True, exist_ok=True)
            for second in range(0, duration, interval_seconds):
                frame = clip.get_frame(second)
                frame_path = output_path / f"sec_{second:04d}.jpg"
                _save_frame_to_path(frame, frame_path)
                exported_files.append(str(frame_path.resolve()))

        result = {
            "exported_count": len(exported_files),
            "output_directory": str(output_path.resolve()),
            "frames": exported_files,
        }
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        raise ImageProcessingError(f"Error exporting per-second frames: {str(e)}")


@mcp.tool()
def export_frames_for_second(
    video_path: str,
    target_second: int,
    output_dir: str,
) -> str:
    """Export all frames within [target_second, target_second + 1) to `output_dir` for burst analysis."""
    try:
        path = _ensure_video_path(video_path)
        output_path = Path(output_dir)
        exported_files: List[str] = []

        with VideoFileClip(str(path)) as clip:
            fps = clip.fps or 0
            if fps <= 0:
                raise ImageProcessingError("Video FPS is unavailable; cannot sample frames.")
            if target_second < 0 or target_second >= clip.duration:
                raise ImageProcessingError("target_second is outside the video duration.")

            interval_duration = min(1.0, clip.duration - target_second)
            max_frames = max(1, math.floor(interval_duration * fps))

            output_path.mkdir(parents=True, exist_ok=True)
            for frame_idx in range(max_frames):
                frame_timestamp = target_second + frame_idx / fps
                frame = clip.get_frame(frame_timestamp)
                frame_path = output_path / f"sec_{target_second:04d}_frame_{frame_idx:04d}.jpg"
                _save_frame_to_path(frame, frame_path)
                exported_files.append(str(frame_path.resolve()))

        result = {
            "second": target_second,
            "exported_count": len(exported_files),
            "frames": exported_files,
        }
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        raise ImageProcessingError(f"Error exporting frames for second {target_second}: {str(e)}")


@mcp.tool()
def export_frame_at_second(
    video_path: str,
    second: int,
    output_path: str,
) -> str:
    """Save the frame located exactly at `second` to `output_path` and return the saved path."""
    try:
        path = _ensure_video_path(video_path)
        target_path = Path(output_path)
        with VideoFileClip(str(path)) as clip:
            if second < 0 or second >= clip.duration:
                raise ImageProcessingError("second is outside the video duration.")
            frame = clip.get_frame(second)
            _save_frame_to_path(frame, target_path)
        return str(target_path.resolve())
    except Exception as e:
        raise ImageProcessingError(f"Error exporting frame at second {second}: {str(e)}")


@mcp.tool()
def export_frame_by_index(
    video_path: str,
    frame_index: int,
    output_path: str,
) -> str:
    """Save the frame referenced by absolute `frame_index` (0-based) to `output_path`."""
    try:
        if frame_index < 0:
            raise ImageProcessingError("frame_index must be non-negative.")

        path = _ensure_video_path(video_path)
        target_path = Path(output_path)

        with VideoFileClip(str(path)) as clip:
            fps = clip.fps or 0
            if fps <= 0:
                raise ImageProcessingError("Video FPS is unavailable; cannot sample by frame index.")
            total_frames = int(clip.duration * fps)
            if frame_index >= total_frames:
                raise ImageProcessingError(f"frame_index must be < {total_frames}.")

            frame_timestamp = frame_index / fps
            frame = clip.get_frame(frame_timestamp)
            _save_frame_to_path(frame, target_path)

        return str(target_path.resolve())
    except Exception as e:
        raise ImageProcessingError(f"Error exporting frame #{frame_index}: {str(e)}")


if __name__ == "__main__":
    mcp.run()
